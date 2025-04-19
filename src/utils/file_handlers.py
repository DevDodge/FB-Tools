import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

class ExcelHandler:
    @staticmethod
    def save_to_excel(data, filename="data/accounts.xlsx"):  # Removed self parameter
        if not os.path.exists(os.path.dirname(filename)):
            os.makedirs(os.path.dirname(filename))

        if not os.path.exists(filename):
            wb = Workbook()
            ws = wb.active
            ws.append(["Name", "Email", "Password", "DOB", "Gender", "Proxy", "Created At"])
        else:
            wb = load_workbook(filename)
            ws = wb.active

        ws.append([
            data['name'],
            data['email'],
            data['password'],
            data['dob'],
            data['gender'],
            data.get('proxy', ''),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ])
        wb.save(filename)
