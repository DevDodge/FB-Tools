# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, messagebox
import csv
import os
import time
import random
import string
import threading
from datetime import datetime
from selenium import webdriver 
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
import json

class KagTools:
    def __init__(self, root):
        self.root = root
        self.root.title("KAG TOOLS")
        self.root.geometry("750x680")
        self.root.resizable(False, False)
        self.entries = {}
        self.setup_ui()

    def setup_ui(self):
        main_frame = tk.Frame(self.root, bg="#1a1a1a")
        main_frame.pack(fill=tk.BOTH, expand=True)

        title_frame = tk.Frame(main_frame, bg="#1a1a1a")
        title_frame.pack(pady=15)

        title = tk.Label(
            title_frame,
            text="KAG TOOLS - Facebook Account Creator",
            font=("Impact", 22, "bold"),
            fg="#FFD700",
            bg="#1a1a1a",
            padx=15,
            pady=8
        )
        title.pack()

        tk.Frame(title_frame, height=3, bg="#FFD700", width=450).pack(pady=5)

        random_btn = tk.Button(
            main_frame,
            text="🎲 توليد بيانات عشوائية",
            command=self.generate_random_data,
            bg="#4CAF50",
            fg="white",
            font=("Arial", 12, "bold"),
            padx=15,
            pady=8,
            cursor="hand2"
        )
        random_btn.pack(pady=10)

        form_frame = tk.Frame(main_frame, bg="#34495e")
        form_frame.pack(pady=10, padx=25, fill=tk.X)

        fields = [
            ("الاسم الكامل", "name"),
            ("البريد الإلكتروني", "email"),
            ("كلمة المرور", "password"),
            ("عدد الحسابات", "account_count"),
            ("البروكسي (اختياري)", "proxy"),
        ]

        for text, field in fields:
            self.create_form_row(form_frame, text, field)

        dob_frame = tk.Frame(form_frame, bg="#34495e")
        dob_frame.pack(pady=5, fill=tk.X)

        tk.Label(
            dob_frame,
            text="تاريخ الميلاد",
            width=15,
            anchor='w',
            bg="#34495e",
            fg="white"
        ).pack(side=tk.LEFT)

        self.day_combo = ttk.Combobox(dob_frame, width=5, values=[str(i) for i in range(1, 32)])
        self.day_combo.pack(side=tk.LEFT, padx=2)

        self.month_combo = ttk.Combobox(dob_frame, width=10, values=[
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"
        ])
        self.month_combo.pack(side=tk.LEFT, padx=2)

        self.year_combo = ttk.Combobox(dob_frame, width=5, values=[str(i) for i in range(1950, 2024)])
        self.year_combo.pack(side=tk.LEFT, padx=2)

        gender_frame = tk.Frame(form_frame, bg="#34495e")
        gender_frame.pack(pady=5, fill=tk.X)

        tk.Label(
            gender_frame,
            text="الجنس",
            width=15,
            anchor='w',
            bg="#34495e",
            fg="white"
        ).pack(side=tk.LEFT)

        self.gender_var = tk.StringVar(value="male")
        genders = [("ذكر", "male"), ("أنثى", "female"), ("مخصص", "custom")]
        for text, val in genders:
            tk.Radiobutton(
                gender_frame,
                text=text,
                variable=self.gender_var,
                value=val,
                bg="#34495e",
                fg="white",
                selectcolor="#2c3e50"
            ).pack(side=tk.LEFT, padx=10)

        self.headless_var = tk.BooleanVar()
        headless_check = tk.Checkbutton(
            main_frame,
            text="تشغيل المتصفح في الخلفية (بدون عرض)",
            variable=self.headless_var,
            bg="#1a1a1a",
            fg="white",
            selectcolor="#2c3e50"
        )
        headless_check.pack(pady=5)

        self.progress = ttk.Progressbar(main_frame, orient='horizontal', length=300, mode='determinate')
        self.progress.pack(pady=10)

        self.status_label = tk.Label(main_frame, text="", bg="#1a1a1a", fg="white")
        self.status_label.pack()

        self.detailed_progress_label = tk.Label(main_frame, text="جاري إنشاء الحسابات...", bg="#1a1a1a", fg="white")
        self.detailed_progress_label.pack(pady=5)

        create_btn = tk.Button(
            main_frame,
            text="🛠️ إنشاء حسابات",
            command=self.validate_form,
            bg="#FF6B6B",
            fg="white",
            font=("Arial", 14, "bold"),
            padx=25,
            pady=12,
            cursor="hand2"
        )
        create_btn.pack(pady=15)

        footer = tk.Label(
            main_frame,
            text="تم التطوير بواسطة KAG Tools",
            bg="#1a1a1a",
            fg="#FFFFFF"
        )
        footer.pack(side=tk.BOTTOM, pady=10)

    def create_form_row(self, parent, label_text, field_key):
        row = tk.Frame(parent, bg="#34495e")
        row.pack(pady=5, fill=tk.X)

        label = tk.Label(row, text=label_text, width=15, anchor='w', bg="#34495e", fg="white")
        label.pack(side=tk.LEFT)

        entry = tk.Entry(row)
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.entries[field_key] = entry

    def generate_random_string(self, length=6):
        return ''.join(random.choices(string.ascii_lowercase + string.digits, k=length))

    def generate_random_data(self):
        first_names_en = ["John", "Michael", "Sarah", "Emma", "James", "Olivia"]
        last_names_en = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia"]
        domains = ["gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "mail.com", "protonmail.com"]
        
        random_name = f"{random.choice(first_names_en)} {random.choice(last_names_en)}"
        random_domain = random.choice(domains)
        random_email = f"{random_name.split()[0].lower()}.{random.choice(last_names_en).lower()}{random.randint(1,99)}@{random_domain}"
        random_password = f"Pass{random.randint(1000,9999)}"

        self.entries['name'].delete(0, tk.END)
        self.entries['name'].insert(0, random_name)
        self.entries['email'].delete(0, tk.END)
        self.entries['email'].insert(0, random_email)
        self.entries['password'].delete(0, tk.END)
        self.entries['password'].insert(0, random_password)
        self.entries['account_count'].delete(0, tk.END)
        self.entries['account_count'].insert(0, str(random.randint(1, 3)))
        self.entries['proxy'].delete(0, tk.END)

        self.day_combo.set(str(random.randint(1, 28)))
        self.month_combo.set(random.choice(self.month_combo['values']))
        self.year_combo.set(str(random.randint(1980, 2000)))
        self.gender_var.set(random.choice(["male", "female"]))

    def validate_form(self):
        threading.Thread(target=self._create_accounts_process).start()

    def _create_accounts_process(self):
        try:
            num_accounts = int(self.entries['account_count'].get().strip())
            self.progress['maximum'] = num_accounts

            for i in range(1, num_accounts + 1):
                try:
                    self.generate_random_data()
                    data = self.prepare_account_data()
                    self.create_account(data)
                    self.save_to_excel(data)
                    self.update_status(f"تم إنشاء الحساب {i}", "#4CAF50")
                    self.update_detailed_progress(f"الحساب {i}/{num_accounts} مكتمل")
                except Exception as e:
                    self.update_status(f"فشل في الحساب {i}: {str(e)}", "red")

                self.progress['value'] = i
                self.root.update()
                time.sleep(random.uniform(2, 4))

            messagebox.showinfo("الانتهاء", "تمت العملية")
        except Exception as e:
            messagebox.showerror("خطأ", str(e))

    def update_status(self, message, color):
        self.status_label.config(text=message, fg=color)
        self.root.update()

    def update_detailed_progress(self, message):
        self.detailed_progress_label.config(text=message)
        self.root.update()

    def prepare_account_data(self):
        return {
            "name": self.entries['name'].get(),
            "email": self.entries['email'].get(),
            "password": self.entries['password'].get(),
            "dob": f"{self.day_combo.get()} {self.month_combo.get()} {self.year_combo.get()}",
            "gender": self.gender_var.get(),
            "proxy": self.entries['proxy'].get(),
            "headless": self.headless_var.get()
        }

    def save_to_excel(self, data):
        filename = "accounts.xlsx"
        if not os.path.exists(filename):
            wb = Workbook()
            ws = wb.active
            ws.append(["Name", "Email", "Password", "DOB", "Gender", "Proxy", "Created At"])
        else:
            wb = load_workbook(filename)
            ws = wb.active

        ws.append([
            data['name'],
            data['email'],
            data['password'],
            data['dob'],
            data['gender'],
            data['proxy'],
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ])
        wb.save(filename)

    def create_account(self, data):
        driver = None
        try:
            user_agents = [
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Safari/605.1.15"
            ]

            chrome_options = Options()
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_argument(f"user-agent={random.choice(user_agents)}")
            if data.get('headless'):
                chrome_options.add_argument("--headless")
            if data.get('proxy'):
                chrome_options.add_argument(f'--proxy-server={data["proxy"]}')

            service = Service(executable_path=os.path.abspath("chromedriver.exe"))
            driver = webdriver.Chrome(service=service, options=chrome_options)
            driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
                "source": """
                    Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
                    const getParameter = WebGLRenderingContext.prototype.getParameter;
                    WebGLRenderingContext.prototype.getParameter = function(parameter) {
                        if (parameter === 37445) return 'Intel Inc.';
                        if (parameter === 37446) return 'Intel Iris OpenGL Engine';
                        return getParameter(parameter);
                    };
                    const toDataURL = HTMLCanvasElement.prototype.toDataURL;
                    HTMLCanvasElement.prototype.toDataURL = function() {
                        return "data:image/png;base64,fake_canvas_data";
                    };
                """
            })

            driver.get("https://www.facebook.com/r.php")

            def slow_type(element, text):
                for char in text:
                    element.send_keys(char)
                    time.sleep(random.uniform(0.05, 0.2))

            slow_type(driver.find_element(By.NAME, "firstname"), data['name'].split()[0])
            slow_type(driver.find_element(By.NAME, "lastname"), data['name'].split()[1])
            slow_type(driver.find_element(By.NAME, "reg_email__"), data['email'])
            time.sleep(1)
            slow_type(driver.find_element(By.NAME, "reg_passwd__"), data['password'])

            driver.find_element(By.ID, "day").send_keys(data['dob'].split()[0])
            driver.find_element(By.ID, "month").send_keys(data['dob'].split()[1])
            driver.find_element(By.ID, "year").send_keys(data['dob'].split()[2])

            gender_element = driver.find_elements(By.NAME, "sex")
            if data['gender'] == "male":
                gender_element[0].click()
            elif data['gender'] == "female":
                gender_element[1].click()

            time.sleep(3)
            driver.find_element(By.NAME, "websubmit").click()
            time.sleep(5)

            cookies = driver.get_cookies()
            email_username = data['email'].split("@")[0]
            with open(f"cookies/{email_username}_cookies.json", "w") as f:
                json.dump(cookies, f)

        except Exception as e:
            raise Exception(f"فشل إنشاء الحساب: {e}")
        finally:
            if driver:
                driver.quit()

    def on_close(self):
        self.root.destroy()

if __name__ == "__main__":
    if not os.path.exists("cookies"):
        os.mkdir("cookies")

    root = tk.Tk()
    app = KagTools(root)
    root.protocol("WM_DELETE_WINDOW", app.on_close)
    root.mainloop()